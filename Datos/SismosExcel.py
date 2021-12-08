from logging import disable
import sys
sys.path.append('D:\\UPT\\2021-II\\Inteligencia de Negocios\\Proyecto')

from BD.sql import clsSQL
import datetime

sql = clsSQL()

def fechasql(cadenafecha):
    dia = int(cadenafecha[0:2])
    mes = int(cadenafecha[3:5])
    year = int(cadenafecha[6:10])
    date = datetime.datetime(year=year, month=mes, day=dia)
    return date.date()

def horasql(cadenahora):
    hora = int(cadenahora[0:2])
    minuto = int(cadenahora[3:5])
    segundo = int(cadenahora[6:8])
    time = datetime.time(hour=hora,minute=minuto,second=segundo)
    return time

import openpyxl

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook(".\\Archivos\\IGP_datos_sismicos.xlsx")

# Define variable to read the active sheet:
worksheet = wookbook.active

# Iterate the loop to read the cell values
for i in range(1, worksheet.max_row):
    contador = 0
    sismo = []
    for col in worksheet.iter_cols(1, worksheet.max_column):
        sismo.append(col[i].value)
        contador += 1
        if contador == 6:
            fecha = str(sismo[0])
            fecha = fechasql(fecha)
            hora = str(sismo[1])
            hora = horasql(hora)
            magnitud = float(sismo[5])
            profundidad = float(sismo[4])
            latitud = float(sismo[2])
            longitud = float(sismo[3])
            query = ("INSERT INTO Sismo ( Magnitud, Fecha, Hora, Latitud, Longitud, Profundidad) VALUES ({0}, '{1}', '{2}', {3}, {4}, {5});"
                    .format(magnitud, fecha, hora, latitud, longitud, profundidad))
            sql.sqlInsert(query)